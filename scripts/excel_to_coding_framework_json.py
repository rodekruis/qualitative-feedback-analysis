#!/usr/bin/env python3
"""Excel (Feedback Tool 23 layout) → nested ``coding_frames`` JSON.

Columns: CODING FRAME, TYPE, CATEGORY, CODE, CODE DESCRIPTION, SENSITIVE, EXAMPLE.
Skips sheets PIVOT and ALL.

    uv run python scripts/excel_to_coding_framework_json.py workbook.xlsx -o out.json
"""

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

SKIP_SHEETS = {"PIVOT", "ALL"}
HEADER = (
    "CODING FRAME",
    "TYPE",
    "CATEGORY",
    "CODE",
    "CODE DESCRIPTION",
    "SENSITIVE",
    "EXAMPLE",
)


def slug(value: object) -> str:
    """Return a lowercase hyphenated slug for stable ``code_id`` segments."""
    normalized = re.sub(r"[^a-z0-9]+", "-", str(value or "").strip().lower()).strip("-")
    return normalized or "x"


def new_code_id(
    coding_frame: str,
    type_name: str,
    category_name: str,
    code_name: str,
    used_code_ids: set[str],
) -> str:
    """Return a unique ``code_id`` from hierarchy labels, suffixing duplicates."""
    base = (
        f"{slug(coding_frame)}:{slug(type_name)}:"
        f"{slug(category_name)}:{slug(code_name)}"
    )
    suffix = 0
    candidate_id = base
    while candidate_id in used_code_ids:
        suffix += 1
        candidate_id = f"{base}:{suffix}"
    used_code_ids.add(candidate_id)
    return candidate_id


def examples_cell(cell_value: object) -> list[str]:
    """Parse the EXAMPLE cell into non-empty example strings."""
    if cell_value is None or not str(cell_value).strip():
        return []
    text = str(cell_value).strip()
    chunks = [
        paragraph.strip() for paragraph in text.split("\n\n") if paragraph.strip()
    ]
    if len(chunks) == 1 and "\n" in chunks[0]:
        chunks = [line.strip() for line in chunks[0].split("\n") if line.strip()]
    return chunks


def sensitive_cell(cell_value: object) -> bool:
    """Return True when the SENSITIVE column indicates a sensitive code."""
    return str(cell_value or "").strip().lower() in {"yes", "y", "true", "1"}


def row_dict(header: tuple[str, ...], row: tuple) -> dict[str, object]:
    """Map column names to row values, padding with None when the row is short."""
    cells = list(row[: len(header)])
    cells += [None] * (len(header) - len(cells))
    return dict(zip(header, cells, strict=True))


def workbook_to_tree(workbook_path: Path) -> dict:
    """Read valid sheets and return the nested ``coding_frames`` payload."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    used_code_ids: set[str] = set()
    # dict key order = first-seen frame / type / category (sheet + row order).
    coding_hierarchy: dict = {}

    try:
        for sheet_name in workbook.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue
            worksheet = workbook[sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            if not header_row:
                continue
            column_names = tuple(str(cell or "").strip() for cell in header_row)
            if column_names[: len(HEADER)] != HEADER:
                continue

            for row in rows:
                if not row or all(cell is None for cell in row):
                    continue
                row_by_column = row_dict(column_names, row)
                coding_frame = row_by_column.get("CODING FRAME")
                type_name = row_by_column.get("TYPE")
                category_name = row_by_column.get("CATEGORY")
                code_name = row_by_column.get("CODE")
                if not all((coding_frame, type_name, category_name, code_name)):
                    continue

                description_cell = row_by_column.get("CODE DESCRIPTION")
                if description_cell is None or not str(description_cell).strip():
                    code_description = None
                else:
                    code_description = str(description_cell).strip()

                code_record = {
                    "code_id": new_code_id(
                        str(coding_frame),
                        str(type_name),
                        str(category_name),
                        str(code_name),
                        used_code_ids,
                    ),
                    "name": str(code_name).strip(),
                    "code_description": code_description,
                    "sensitive": sensitive_cell(row_by_column.get("SENSITIVE")),
                    "examples": examples_cell(row_by_column.get("EXAMPLE")),
                }
                coding_frame_key = str(coding_frame).strip()
                type_key = str(type_name).strip()
                category_key = str(category_name).strip()
                coding_hierarchy.setdefault(coding_frame_key, {}).setdefault(
                    type_key, {}
                ).setdefault(category_key, []).append(code_record)
    finally:
        workbook.close()

    coding_frames = [
        {
            "name": coding_frame_name,
            "types": [
                {
                    "name": type_name,
                    "categories": [
                        {"name": category_name, "codes": code_list}
                        for category_name, code_list in categories_to_codes.items()
                    ],
                }
                for type_name, categories_to_codes in types_to_categories.items()
            ],
        }
        for coding_frame_name, types_to_categories in coding_hierarchy.items()
    ]
    return {"coding_frames": coding_frames}


def main() -> int:
    """Run the CLI: parse arguments, convert the workbook, write JSON."""
    argument_parser = argparse.ArgumentParser(
        description="Convert coding framework Excel to JSON."
    )
    argument_parser.add_argument("workbook", type=Path, help="Input .xlsx path")
    argument_parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Output .json (default: <workbook_stem>_coding_framework.json beside input)",
    )
    arguments = argument_parser.parse_args()
    if not arguments.workbook.is_file():
        print(f"Not found: {arguments.workbook}", file=sys.stderr)
        return 1

    output_path = arguments.output or arguments.workbook.with_name(
        f"{arguments.workbook.stem}_coding_framework.json"
    )
    payload = workbook_to_tree(arguments.workbook)
    output_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
